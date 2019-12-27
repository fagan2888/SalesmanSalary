from openpyxl import load_workbook
import logging
import sqlite3

logging.disable(logging.NOTSET)
logging.basicConfig(level=logging.DEBUG,
                    format=' %(asctime)s | %(message)s')


def update(conn, cur, file_name):

    cur.execute(f"DELETE FROM '{file_name}'")

    wb = load_workbook(f'司龄工资\\{file_name}.xlsx', read_only=True)

    if '12个月' in file_name:
        ws = wb['page']
        begin_row = 2
    else:
        ws = wb['司龄工资核定表']
        begin_row = 3

    for row in ws.iter_rows(min_row=begin_row,
                            max_row=ws.max_row,
                            min_col=1,
                            max_col=ws.max_column):
        str_sql = f"INSERT INTO '{file_name}' VALUES ("
        for r in row:
            str_sql += f"'{r.value}', "

        str_sql = str_sql[:-2] + ')'
        cur.execute(str_sql)

    conn.commit()

    logging.debug(f"{file_name} 表更新完成")


conn = sqlite3.connect(r'Data\data.db')
cur = conn.cursor()

update(conn, cur, "司龄工资核定表")
update(conn, cur, "前线人员信息表")
update(conn, cur, '滚动12个月签单保费')
update(conn, cur, '滚动12个月同期签单保费')
