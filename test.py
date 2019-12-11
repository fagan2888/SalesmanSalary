from openpyxl import load_workbook
import sqlite3


def update(conn, cur, file_name):
    wb = load_workbook(f'{file_name}.xlsx', read_only=True)
    ws = wb['page1']

    str_sql = f"CREATE TABLE '{file_name}' ("

    if file_name == '前线':
        title_row = 1
    else:
        title_row = 2

    for row in ws.iter_rows(min_row=title_row,
                            max_row=title_row,
                            min_col=1,
                            max_col=ws.max_column):
        for r in row:
            str_sql += f"'{r.value[:10]}' TEXT" + ', '

        str_sql = str_sql[:-2] + ')'

    cur.execute(str_sql)

    for row in ws.iter_rows(min_row=3,
                            max_row=ws.max_row,
                            min_col=1,
                            max_col=ws.max_column):
        str_sql = f"INSERT INTO '{file_name}' VALUES ("
        for r in row:
            str_sql += f"'{r.value}', "

        str_sql = str_sql[:-2] + ')'
        cur.execute(str_sql)

    conn.commit()


conn = sqlite3.connect('data.db')
cur = conn.cursor()

update(conn, cur, "司龄工资")
update(conn, cur, "前线")
