from openpyxl import Workbook
from datetime import date
from openpyxl.styles import (Border,
                             Font,
                             NamedStyle,
                             PatternFill,
                             Side,
                             Alignment)


def write_excel(rens):
    wb = Workbook()
    ws = wb.active
    ws.title = "司龄工资核定表"

    year = int(date.today().strftime('%Y'))
    month = int(date.today().strftime('%m'))

    font_bold = Font(name='微软雅黑',
                     size=12,
                     bold=True)
    font = Font(name='微软雅黑',
                size=12)
    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000'))
    fill = PatternFill(fill_type=None,
                       bgColor='CCCCCC')
    alignment = Alignment(horizontal='center',
                          vertical='center',
                          wrap_text=True)
    style_bold = NamedStyle(name='style_bold',
                            font=font_bold,
                            border=border,
                            alignment=alignment)
    style_no_fill = NamedStyle(name='sytle_no_fill',
                               font=font,
                               border=border,
                               alignment=alignment)
    style_fill = NamedStyle(name='style_fill',
                            font=font,
                            border=border,
                            fill=fill,
                            alignment=alignment)

    wb.add_named_style(style_bold)
    wb.add_named_style(style_no_fill)
    wb.add_named_style(style_fill)

    row = ['司龄工资核定表']
    ws.append(row)

    row = ['序号',
           '中支',
           '机构',
           '姓名',
           '合同类型',
           '现任职级',
           '入司职级',
           '入司时间',
           '入司月份',
           '是否重新核定',
           f'滚动12个月签单保费\n({year-1}-{month-1}至{year}-{month-2})',
           f'滚动12个月同期签单保费\n({year-2}-{month-1}至{year-1}-{month-2})',
           '同比增长率',
           '现任司龄工资',
           '司龄工资变化',
           '考核后司龄工资',
           '核定说明']

    ws.append(row)

    i = 1

    for ren in rens:
       row = [i,
              ren.zhong_zhi,
              ren.ji_gou,
              ren.name,
              ren.he_tong,
              ren.zhi_ji,
              ren.ru_si_shi_jian,
              ]

    ws.merge_cells('A1:Q1')
    ws['A1'].style = style_bold


    # nrow = 1
    # ws.cell(row=nrow, column=1).value = '前线人员司龄工资核定表'
    # ws.cell(row=nrow, column=1).font = font_bold
    # ws.cell(row=nrow, column=1).alignment = alignment
    # ws.merge_cells('A1:O1')

    # nrow += 1
    # ws.cell(row=nrow, column=1).value = '序号'
    # ws.cell(row=nrow, column=2).value = '机构'
    # ws.cell(row=nrow, column=3).value = '团队'
    # ws.cell(row=nrow, column=4).value = '姓名'
    # ws.cell(row=nrow, column=5).value = '合同类型'
    # ws.cell(row=nrow, column=6).value = '现任职级'
    # ws.cell(row=nrow, column=7).value = '入司职级'
    # ws.cell(row=nrow, column=8).value = '入司时间'
    # ws.cell(row=nrow, column=9).value = '滚动12个月签单保费'
    # ws.cell(row=nrow, column=10).value = '滚动12个月同期签单保费'
    # ws.cell(row=nrow, column=11).value = '同比增长率'
    # ws.cell(row=nrow, column=12).value = '现任司龄工资'
    # ws.cell(row=nrow, column=13).value = '司龄工资变化'
    # ws.cell(row=nrow, column=14).value = '考核后司龄工资'
    # ws.cell(row=nrow, column=15).value = '核定说明'

    # ws.column_dimensions['A'].width = 4
    # ws.column_dimensions['B'].width = 6
    # ws.column_dimensions['C'].width = 20
    # ws.column_dimensions['D'].width = 9
    # ws.column_dimensions['E'].width = 6
    # ws.column_dimensions['F'].width = 20
    # ws.column_dimensions['G'].width = 20
    # ws.column_dimensions['H'].width = 15
    # ws.column_dimensions['I'].width = 30
    # ws.column_dimensions['J'].width = 30
    # ws.column_dimensions['K'].width = 30
    # ws.column_dimensions['L'].width = 10
    # ws.column_dimensions['M'].width = 10
    # ws.column_dimensions['N'].width = 10
    # ws.column_dimensions['O'].width = 70

    # for i in range(1, 16):
    #     ws.cell(row=nrow, column=i).style = style_bold

    # nrow += 1

    # for ren in info:
    #     ws.cell(row=nrow, column=1).value = nrow - 2
    #     ws.cell(row=nrow, column=2).value = ren.zhong_zhi
    #     ws.cell(row=nrow, column=3).value = ren.tuan_dui
    #     ws.cell(row=nrow, column=4).value = ren.xing_ming
    #     ws.cell(row=nrow, column=5).value = ren.he_tong_lei_xing
    #     ws.cell(row=nrow, column=6).value = ren.zhi_ji
    #     ws.cell(row=nrow, column=7).value = ren.ru_si_zhi_ji
    #     ws.cell(row=nrow, column=8).value = ren.ru_si_shi_jian
    #     ws.cell(row=nrow, column=9).value = ren.shang_nian_bao_fei
    #     ws.cell(row=nrow, column=10).value = ren.tong_qi_bao_fei
    #     ws.cell(row=nrow, column=11).value = ren.tong_bi
    #     ws.cell(row=nrow, column=12).value = ren.yuan_si_ling_gong_zi
    #     ws.cell(row=nrow, column=13).value = ren.si_ling_gong_zi_bian_hua
    #     ws.cell(row=nrow, column=14).value = ren.xin_si_ling_gong_zi
    #     ws.cell(row=nrow, column=15).value = ren.he_ding_shuo_ming
    #     if nrow % 2 == 1:
    #         style = style_fill
    #     else:
    #         style = style_no_fill

    #     for i in range(1, 16):
    #         ws.cell(row=nrow, column=i).style = style

    #     nrow += 1

    wb.save('司龄工资核定表.xlsx')
